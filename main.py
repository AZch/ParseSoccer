import datetime

from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from ExecReq import *
from openpyxl import *

ligs = [
        'r[43]', 'r[44]', 'r[8]', 'r[1]', 'r[10]', 'r[11]', 'r[2]', 'r[45]', 'r[13]',
        'r[95]', 'r[15]', 'r[16]', 'r[17]', 'r[18]', 'r[19]', 'r[47]', 'r[48]', 'r[96]',
        'r[51]', 'r[20]', 'r[52]', 'r[21]', 'r[22]', 'r[53]', 'r[54]', 'r[58]', 'r[59]',
        'r[61]', 'r[99]', 'r[24]', 'r[62]', 'r[25]', 'r[26]', 'r[27]', 'r[110]', 'r[6]',
        'r[30]', 'r[31]', 'r[32]', 'r[33]', 'r[34]', 'r[35]', 'r[36]', 'r[67]', 'r[38]',
        'r[39]', 'r[40]', 'r[41]', 'r[69]', 'r[70]', 'r[71]', 'r[73]', 'r[42]', 'r[102]'
        ]

def main():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'League'
    ws['B1'] = 'Home'
    ws['C1'] = 'Away'
    ws['D1'] = 'Score'
    ws['E1'] = 'Odds 1'
    ws['F1'] = 'Odds X'
    ws['G1'] = 'Odds 2'
    ws['H1'] = 'Ver 1'
    ws['I1'] = 'Ver X'
    ws['J1'] = 'Ver 2'
    ws['K1'] = 'Total'
    ws['L1'] = 'Under'
    ws['M1'] = 'Over'
    ws['N1'] = 'U ver'
    ws['O1'] = 'O ver'
    ws['P1'] = 'date'
    chrome_options = Options()
    notFin = True
    while notFin:
        try:
            driver = webdriver.Chrome("/home/az/ProjectsData/Drivers/chromedriver",chrome_options=chrome_options)
            driver.set_page_load_timeout(40)
            driver.get('https://alexbetting.com/auth/')
            notFin = False
        except:
            print('TIMEOUT:\n', traceback.format_exc())
            print('timeout')
            driver.close()
            time.sleep(1)
    login(driver)
    date = datetime.date(2017, 1, 1)
    while len(driver.window_handles) > 1:
        driver.switch_to.window(driver.window_handles[1])
        driver.close()
        driver.switch_to.window(driver.window_handles[0])

    currStr = 2
    saveCurrStr = 1
    while not (date.day == 31 and date.month == 12 and date.year == 2018):
        if (saveCurrStr == currStr):
            date = date + datetime.timedelta(days=-5)

        while True:
            try:
                clickGetElemEnter(driver, "//a[@href='/matches/statp/']")
                break
            except:
                print('TIMEOUT:\n', traceback.format_exc())
                print('timeout')
                time.sleep(3)
        print(str(date))
        for lig in ligs:
            clickGetElemSpace(driver, "//input[@name='" + lig + "']")
        if len(str(date.day)) == 1:
            clickGetElemSimple(driver, "//select[@name='day']/option[@value='0" + str(date.day) + "']")
        else:
            clickGetElemSimple(driver, "//select[@name='day']/option[@value='" + str(date.day) + "']")
        if len(str(date.month)) == 1:
            clickGetElemSimple(driver, "//select[@name='month']/option[@value='0" + str(date.month) + "']")
        else:
            clickGetElemSimple(driver, "//select[@name='month']/option[@value='" + str(date.month) + "']")

        clickGetElemSimple(driver, "//select[@name='year']/option[contains(text(), '" + str(date.year) + "')]")

        clickGetElemSimple(driver, "//select[@name='dodata']/option[@value='4']")
        date = date + datetime.timedelta(days=5)
        saveCurrStr = currStr
        clickGetElemEnter(driver, "//input[@value='Показать']")
        time.sleep(1)
        while True:
            try:
                for elem in getElemsByXPath("//div[@id='all']/div/pre", driver):
                    print(elem.text)
                    currStr = fillExcel(currStr, elem.text, ws)
                if clickGetElemEnter(driver, "//b[contains(text(), '»»»')]/parent::a", maxTimeout = 1) == False:
                    break
            except:
                print('errot:\n', traceback.format_exc())
                driver.refresh()
                time.sleep(1)

        wb.save("report.xlsx")


def fillExcel(currStr, text, ws):
    splitText = text.split('\n')
    ligue = splitText[-1]
    for i in range(len(splitText) - 1):
        if (len(splitText[i].split(' ')) > 1):
            if (len(splitText[i].split(' ')[1].split('.')) == 3):
                oneStr = splitText[i].split()
                twoStr = splitText[i + 1].split()
                ws['A' + str(currStr)] = ligue[:-20]
                teamName = ""
                teamName += oneStr[2]
                idOne = 3
                while len(oneStr[idOne].split(':')) != 2:
                    teamName += " " +  oneStr[idOne]
                    idOne += 1
                ws['B' + str(currStr)] = teamName
                teamName = ""
                teamName += twoStr[2]
                idTwo = 3
                while len(twoStr[idTwo].split(':')) != 2:
                    if (twoStr[idTwo][-1] == '%'):
                        idTwo -= 1
                        break
                    teamName += " " +  twoStr[idTwo]
                    idTwo += 1


                ws['C' + str(currStr)] = teamName
                idOne -= 1
                ws['D' + str(currStr)] = oneStr[idOne + 1]
                ws['E' + str(currStr)] = float(oneStr[idOne + 2])
                ws['F' + str(currStr)] = float(oneStr[idOne + 3])
                ws['G' + str(currStr)] = float(oneStr[idOne + 4])
                ws['H' + str(currStr)] = twoStr[idTwo + 1][:-1]
                ws['I' + str(currStr)] = twoStr[idTwo + 2][:-1]
                ws['J' + str(currStr)] = twoStr[idTwo + 3][:-1]
                try:
                    ws['K' + str(currStr)] = float(oneStr[idOne + 14])
                except:
                    ws['K' + str(currStr)] = float(oneStr[len(oneStr) - 3])
                try:
                    ws['L' + str(currStr)] = float(oneStr[idOne + 15])
                except:
                    ws['L' + str(currStr)] = 0
                try:
                    ws['M' + str(currStr)] = float(oneStr[idOne + 16])
                except:
                    ws['M' + str(currStr)] = 0
                try:
                    ws['N' + str(currStr)] = twoStr[idTwo + 9][:-1]
                except:
                    ws['N' + str(currStr)] = 0
                try:
                    ws['O' + str(currStr)] = twoStr[idTwo + 10][:-1]
                except:
                    ws['O' + str(currStr)] = 0
                ws['P' + str(currStr)] = splitText[i].split(' ')[1]
                currStr += 1
    return currStr

def login(driver, name = '', password = ''):
    time.sleep(1)
    while (True):
        clickGetElemEnter(driver, "//a[contains(text(), 'Авторизация')]")
        isContinue = False
        while len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[1])
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            isContinue = True
        if isContinue:
            continue
        getElemByXPath("//input[@name='login']", driver).send_keys(name)
        getElemByXPath("//input[@name='pass']", driver).send_keys(password)
        while (True):
            time.sleep(3)
            clickGetElemEnter(driver, "//input[@class='but']")
            isContinue = False
            while len(driver.window_handles) > 1:
                driver.switch_to.window(driver.window_handles[1])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                isContinue = True
            if isContinue:
                continue
            break
        break
    driver.refresh()





if __name__ == '__main__':
    main()
